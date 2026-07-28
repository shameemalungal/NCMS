from openpyxl import load_workbook


class ExcelReader:
    REQUIRED_SHEETS = [
        "Pashudhan ID wise Achievement",
        "Panchayath Population",
    ]

    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = None

    def load(self):
        self.workbook = load_workbook(
            self.filepath,
            data_only=True,
        )

    def validate(self):
        missing = []

        for sheet in self.REQUIRED_SHEETS:
            if sheet not in self.workbook.sheetnames:
                missing.append(sheet)

        if missing:
            raise ValueError(
                "Missing worksheet(s): "
                + ", ".join(missing)
            )

    def get_population(self):
        ws = self.workbook["Panchayath Population"]

        data = {}

        for row in ws.iter_rows(min_row=2, values_only=True):

            if not row[0]:
                continue

            data[str(row[0]).strip().upper()] = row[1]

        return data

    def get_squad_rows(self):

        ws = self.workbook["Pashudhan ID wise Achievement"]

        rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):

            if not row[0]:
                continue

            rows.append(
                {
                    "panchayath": str(row[0]).strip().upper(),
                    "squad_no": int(row[1]),
                    "days": int(row[2]),
                    "member": str(row[3]).strip(),
                    "pashudhan_id": (
                        str(row[4]).strip()
                        if row[4]
                        else None
                    ),
                }
            )

        return rows