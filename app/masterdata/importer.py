import re
import time
from pathlib import Path

from openpyxl import load_workbook

from app.extensions import db
from app.masterdata.services import (
    EXPECTED_SHEET,
    POPULATION_SHEET,
    clean_text,
    parse_non_negative_integer,
    validate_excel,
)
from app.models import (
    Campaign,
    ImportHistory,
    Panchayath,
    Squad,
    SquadMember,
)
from app.services.target_service import TargetService


class MasterDataImporter:
    """
    Import one validated workbook into an existing campaign.

    After all Panchayaths and squads are imported,
    Panchayath population is automatically distributed
    among all squads belonging to that Panchayath.
    """

    def __init__(self, file_path: str, campaign_id: int):
        self.file_path = file_path
        self.campaign_id = campaign_id

        self.summary = {
            "source_rows": 0,
            "panchayaths_created": 0,
            "panchayaths_updated": 0,
            "squads_created": 0,
            "squads_updated": 0,
            "members_created": 0,
            "members_updated": 0,
            "targets_recalculated": 0,
            "errors": [],
        }

    # ==========================================================
    # Main Import
    # ==========================================================

    def import_data(self) -> dict:

        started_at = time.monotonic()

        validation = validate_excel(
            self.file_path
        )

        self.summary["source_rows"] = validation.rows

        if not validation.success:

            self.summary["errors"].extend(
                validation.errors
            )

            return self.summary

        campaign = db.session.get(
            Campaign,
            self.campaign_id
        )

        if campaign is None:

            self.summary["errors"].append(
                "Selected campaign not found."
            )

            return self.summary

        try:

            workbook = load_workbook(
                self.file_path,
                data_only=True,
                read_only=True,
            )

            try:

                # ------------------------------------------
                # Read Panchayath population sheet
                # ------------------------------------------

                population_map = self._population_map(
                    workbook
                )

                # ------------------------------------------
                # Main master-data sheet
                # ------------------------------------------

                sheet = workbook[EXPECTED_SHEET]

                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True,
                ):

                    if not any(
                        clean_text(value)
                        for value in row
                    ):
                        continue

                    self._import_row(
                        row,
                        validation.column_indexes,
                        campaign,
                        population_map,
                    )

                # ------------------------------------------
                # Flush imported squads before calculating
                # targets.
                #
                # This ensures TargetService can see every
                # newly imported squad.
                # ------------------------------------------

                db.session.flush()

                # ------------------------------------------
                # AUTOMATIC TARGET DISTRIBUTION
                #
                # Panchayath Population
                # ----------------------
                # Number of Squads
                #
                # TargetService also distributes any
                # remainder so that total squad targets
                # exactly equal Panchayath population.
                # ------------------------------------------

                target_results = (
                    TargetService
                    .recalculate_campaign_targets(
                        campaign.id,
                        commit=False,
                    )
                )

                self.summary[
                    "targets_recalculated"
                ] = sum(
                    result["squad_count"]
                    for result in target_results
                )

                # ------------------------------------------
                # Safety validation
                #
                # Never commit an import if squad targets
                # do not add up to Panchayath population.
                # ------------------------------------------

                for result in target_results:

                    if (
                        result["allocated_total"]
                        != result["population"]
                    ):

                        raise ValueError(
                            "Target allocation mismatch for "
                            f"{result['panchayath']}: "
                            f"population="
                            f"{result['population']}, "
                            f"allocated="
                            f"{result['allocated_total']}."
                        )

                # ------------------------------------------
                # Import history
                # ------------------------------------------

                self._add_history(
                    campaign=campaign,
                    status="SUCCESS",
                    duration_seconds=round(
                        time.monotonic()
                        - started_at
                    ),
                )

                # ------------------------------------------
                # ONE COMMIT
                #
                # Master data + calculated squad targets
                # are saved together.
                # ------------------------------------------

                db.session.commit()

                return self.summary

            finally:

                workbook.close()

        except Exception as ex:

            db.session.rollback()

            self.summary["errors"].append(
                f"Import failed: {ex}"
            )

            self._record_failed_import(
                campaign,
                started_at,
            )

            return self.summary

    # ==========================================================
    # Population Map
    # ==========================================================

    def _population_map(
        self,
        workbook,
    ) -> dict[str, int]:

        if POPULATION_SHEET not in workbook.sheetnames:
            return {}

        population_map = {}

        sheet = workbook[POPULATION_SHEET]

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):

            if not row:
                continue

            name = clean_text(
                row[0]
                if len(row) > 0
                else None
            )

            population = (
                parse_non_negative_integer(
                    row[1]
                    if len(row) > 1
                    else None
                )
            )

            if (
                not name
                or name.lower() == "nan"
                or "total" in name.lower()
            ):
                continue

            population_map[
                name.casefold()
            ] = population or 0

        return population_map

    # ==========================================================
    # Import Individual Row
    # ==========================================================

    def _import_row(
        self,
        row,
        column_indexes,
        campaign,
        population_map,
    ) -> None:

        # ------------------------------------------
        # Read values
        # ------------------------------------------

        panchayath_name = clean_text(
            row[
                column_indexes["PANCHAYATH"]
            ]
        )

        squad_no = (
            parse_non_negative_integer(
                row[
                    column_indexes["SQUAD No."]
                ]
            )
        )

        squad_days = (
            parse_non_negative_integer(
                row[
                    column_indexes["SQUAD DAYS"]
                ]
            )
        )

        member_text = clean_text(
            row[
                column_indexes["SQUAD"]
            ]
        )

        pashudhan_id = clean_text(
            row[
                column_indexes["Pashudhan ID"]
            ]
        )

        population = population_map.get(
            panchayath_name.casefold(),
            0,
        )

        # ==================================================
        # Panchayath
        # ==================================================

        panchayath = (
            Panchayath.query
            .filter_by(
                campaign_id=campaign.id,
                name=panchayath_name,
            )
            .first()
        )

        if panchayath is None:

            panchayath = Panchayath(
                campaign_id=campaign.id,
                name=panchayath_name,
                population=population,
            )

            db.session.add(
                panchayath
            )

            db.session.flush()

            self.summary[
                "panchayaths_created"
            ] += 1

        elif panchayath.population != population:

            panchayath.population = population

            self.summary[
                "panchayaths_updated"
            ] += 1

        # ==================================================
        # Squad
        # ==================================================

        squad = (
            Squad.query
            .filter_by(
                campaign_id=campaign.id,
                squad_no=squad_no,
            )
            .first()
        )

        if squad is None:

            # Target is temporarily zero.
            #
            # The correct target will be calculated
            # automatically after ALL squads have been
            # imported.

            squad = Squad(
                campaign_id=campaign.id,
                panchayath_id=panchayath.id,
                squad_no=squad_no,
                squad_days=squad_days,
                target=0,
                status="Pending",
            )

            db.session.add(
                squad
            )

            db.session.flush()

            self.summary[
                "squads_created"
            ] += 1

        else:

            changed = (
                squad.panchayath_id
                != panchayath.id
                or squad.squad_days
                != squad_days
            )

            squad.panchayath_id = (
                panchayath.id
            )

            squad.squad_days = (
                squad_days
            )

            # Do NOT set squad.target here.
            #
            # TargetService will calculate it after
            # the entire workbook has been imported.

            if changed:

                self.summary[
                    "squads_updated"
                ] += 1

        # ==================================================
        # Squad Member
        # ==================================================

        member_name, office = (
            self._split_member(
                member_text
            )
        )

        member = (
            SquadMember.query
            .filter_by(
                squad_id=squad.id,
                pashudhan_id=pashudhan_id,
            )
            .first()
        )

        if member is None:

            db.session.add(
                SquadMember(
                    squad_id=squad.id,
                    member_name=member_name,
                    office=office,
                    pashudhan_id=pashudhan_id,
                    full_text=member_text,
                )
            )

            self.summary[
                "members_created"
            ] += 1

        else:

            changed = (
                member.member_name
                != member_name
                or member.office
                != office
                or member.full_text
                != member_text
            )

            member.member_name = (
                member_name
            )

            member.office = (
                office
            )

            member.full_text = (
                member_text
            )

            if changed:

                self.summary[
                    "members_updated"
                ] += 1

    # ==========================================================
    # Member Parser
    # ==========================================================

    @staticmethod
    def _split_member(
        member_text: str,
    ) -> tuple[str, str]:

        match = re.match(
            r"^(.*?)\s*\((.*?)\)$",
            member_text,
        )

        if match:

            return (
                match.group(1).strip(),
                match.group(2).strip(),
            )

        return member_text, ""

    # ==========================================================
    # Import History
    # ==========================================================

    def _add_history(
        self,
        campaign,
        status: str,
        duration_seconds: int,
    ) -> None:

        db.session.add(
            ImportHistory(
                campaign_id=campaign.id,
                import_type="MASTER_DATA",
                filename=Path(
                    self.file_path
                ).name,
                total_rows=(
                    self.summary[
                        "source_rows"
                    ]
                ),
                imported_rows=(
                    self.summary[
                        "source_rows"
                    ]
                ),
                failed_rows=0,
                status=status,
                duration_seconds=(
                    duration_seconds
                ),
            )
        )

    # ==========================================================
    # Failed Import History
    # ==========================================================

    def _record_failed_import(
        self,
        campaign,
        started_at: float,
    ) -> None:

        try:

            db.session.add(
                ImportHistory(
                    campaign_id=campaign.id,
                    import_type="MASTER_DATA",
                    filename=Path(
                        self.file_path
                    ).name,
                    total_rows=(
                        self.summary[
                            "source_rows"
                        ]
                    ),
                    imported_rows=0,
                    failed_rows=(
                        self.summary[
                            "source_rows"
                        ]
                    ),
                    status="FAILED",
                    duration_seconds=round(
                        time.monotonic()
                        - started_at
                    ),
                )
            )

            db.session.commit()

        except Exception:

            db.session.rollback()