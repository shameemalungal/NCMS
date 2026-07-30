from dataclasses import dataclass, field
from numbers import Real
from typing import Any

from openpyxl import load_workbook


EXPECTED_SHEET = "Pashudhan ID wise Achievement"
POPULATION_SHEET = "Panchayath Population"
REQUIRED_COLUMNS = [
    "PANCHAYATH",
    "SQUAD No.",
    "SQUAD DAYS",
    "SQUAD",
    "Pashudhan ID",
]


@dataclass
class ValidationResult:
    success: bool = False
    message: str = ""
    headers: list[str] = field(default_factory=list)
    rows: int = 0
    errors: list[str] = field(default_factory=list)
    preview: list[list[Any]] = field(default_factory=list)
    column_indexes: dict[str, int] = field(default_factory=dict)


def _normalize_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def parse_non_negative_integer(value: Any) -> int | None:
    if isinstance(value, bool) or value is None or clean_text(value) == "":
        return None
    if isinstance(value, Real):
        return int(value) if float(value).is_integer() and value >= 0 else None
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return int(parsed) if parsed.is_integer() and parsed >= 0 else None


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(clean_text(value) == "" for value in row)


def _validate_population_sheet(workbook, result: ValidationResult) -> None:
    if POPULATION_SHEET not in workbook.sheetnames:
        return

    sheet = workbook[POPULATION_SHEET]
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_blank_row(row):
            continue
        name = clean_text(row[0] if len(row) > 0 else None)
        population = parse_non_negative_integer(row[1] if len(row) > 1 else None)
        if not name or name.lower() == "nan" or "total" in name.lower():
            continue
        if population is None:
            result.errors.append(
                f"{POPULATION_SHEET}: row {row_number} has an invalid population for '{name}'."
            )


def validate_excel(filepath: str) -> ValidationResult:
    result = ValidationResult()

    try:
        workbook = load_workbook(filepath, data_only=True, read_only=True)
        try:
            if EXPECTED_SHEET not in workbook.sheetnames:
                result.errors.append(f"Sheet '{EXPECTED_SHEET}' not found.")
                return result

            sheet = workbook[EXPECTED_SHEET]
            raw_headers = [_normalize_header(cell.value) for cell in sheet[1]]
            result.headers = [header for header in raw_headers if header]

            missing = [column for column in REQUIRED_COLUMNS if column not in raw_headers]
            if missing:
                result.errors.append(f"Missing columns: {', '.join(missing)}")
                return result

            result.column_indexes = {
                column: raw_headers.index(column) for column in REQUIRED_COLUMNS
            }
            _validate_population_sheet(workbook, result)

            squad_panchayaths: dict[int, str] = {}
            member_keys: set[tuple[int, str]] = set()

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if _is_blank_row(row):
                    continue

                result.rows += 1
                if len(result.preview) < 20:
                    result.preview.append([
                        row[index] if index < len(row) else None
                        for index, header in enumerate(raw_headers)
                        if header
                    ])

                panchayath = clean_text(row[result.column_indexes["PANCHAYATH"]])
                squad_no = parse_non_negative_integer(row[result.column_indexes["SQUAD No."]])
                squad_days = parse_non_negative_integer(row[result.column_indexes["SQUAD DAYS"]])
                member = clean_text(row[result.column_indexes["SQUAD"]])
                pashudhan_id = clean_text(row[result.column_indexes["Pashudhan ID"]])

                if not panchayath or panchayath.lower() == "nan" or "total" in panchayath.lower():
                    result.errors.append(f"Row {row_number}: PANCHAYATH is required.")
                if squad_no is None or squad_no == 0:
                    result.errors.append(f"Row {row_number}: SQUAD No. must be a positive whole number.")
                if squad_days is None:
                    result.errors.append(f"Row {row_number}: SQUAD DAYS must be a non-negative whole number.")
                if not member or member.lower() == "nan":
                    result.errors.append(f"Row {row_number}: SQUAD member is required.")
                if not pashudhan_id or pashudhan_id.lower() == "nan":
                    result.errors.append(f"Row {row_number}: Pashudhan ID is required.")

                if squad_no is not None and squad_no > 0 and panchayath:
                    existing_panchayath = squad_panchayaths.setdefault(squad_no, panchayath)
                    if existing_panchayath != panchayath:
                        result.errors.append(
                            f"Row {row_number}: squad {squad_no} is assigned to both "
                            f"'{existing_panchayath}' and '{panchayath}'."
                        )

                if squad_no is not None and squad_no > 0 and pashudhan_id:
                    key = (squad_no, pashudhan_id)
                    if key in member_keys:
                        result.errors.append(
                            f"Row {row_number}: duplicate Pashudhan ID '{pashudhan_id}' "
                            f"for squad {squad_no}."
                        )
                    member_keys.add(key)

            if result.rows == 0:
                result.errors.append("The achievement sheet does not contain any data rows.")

            if result.errors:
                return result

            result.success = True
            result.message = "Validation completed successfully."
            return result
        finally:
            workbook.close()
    except Exception as ex:
        result.errors.append(f"Unable to read the Excel file: {ex}")
        return result
