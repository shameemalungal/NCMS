from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


class ReportService:
    """
    Excel report generation service for NCMS.

    Report calculations are not performed here.
    This service consumes data already calculated by
    MonitoringService so that Reports and Squad Monitor
    always use the same figures.
    """

    # ==========================================================
    # Colours
    # ==========================================================

    GREEN = "087F4F"
    LIGHT_GREEN = "E4F7ED"

    DARK = "172033"
    GREY = "667085"

    LIGHT_GREY = "F4F7F9"
    BORDER_GREY = "DDE3E8"

    WHITE = "FFFFFF"

    RED = "C92A3A"
    LIGHT_RED = "FDE7E9"

    AMBER = "9A6700"
    LIGHT_AMBER = "FFF3D6"

    # ==========================================================
    # Panchayath Achievement Export
    # ==========================================================

    @staticmethod
    def create_panchayath_achievement_workbook(
        campaign,
        monitoring,
        summary,
        report_filter="all",
    ):
        """
        Create the Panchayath Achievement Excel workbook.

        Supported filters:

        all
        vaccination-achieved
        pashudhan-achieved
        both-achieved
        low-vaccination
        low-pashudhan
        low-both
        pending
        """

        # ------------------------------------------------------
        # Apply report filter
        # ------------------------------------------------------

        filtered_rows = []

        for row in monitoring:

            if ReportService._matches_filter(
                row,
                report_filter,
            ):
                filtered_rows.append(row)

        # ------------------------------------------------------
        # Workbook
        # ------------------------------------------------------

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Panchayath Achievement"

        # ------------------------------------------------------
        # Page Setup
        # ------------------------------------------------------

        worksheet.sheet_view.showGridLines = False

        worksheet.freeze_panes = "A8"

        worksheet.auto_filter.ref = (
            f"A7:K{7 + len(filtered_rows)}"
        )

        worksheet.page_setup.orientation = "landscape"

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        worksheet.print_title_rows = "1:7"

        # ======================================================
        # TITLE
        # ======================================================

        worksheet.merge_cells(
            "A1:K1"
        )

        title_cell = worksheet["A1"]

        title_cell.value = (
            "NCMS - Panchayath Achievement Report"
        )

        title_cell.font = Font(
            size=16,
            bold=True,
            color=ReportService.WHITE,
        )

        title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=ReportService.GREEN,
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 28

        # ======================================================
        # CAMPAIGN
        # ======================================================

        worksheet.merge_cells(
            "A2:K2"
        )

        campaign_cell = worksheet["A2"]

        campaign_name = (
            campaign.name
            if campaign
            else ""
        )

        campaign_cell.value = (
            f"Campaign: {campaign_name}"
        )

        campaign_cell.font = Font(
            size=11,
            bold=True,
            color=ReportService.DARK,
        )

        campaign_cell.alignment = Alignment(
            horizontal="center",
        )

        # ======================================================
        # REPORT FILTER
        # ======================================================

        worksheet.merge_cells(
            "A3:K3"
        )

        filter_cell = worksheet["A3"]

        filter_cell.value = (
            "Report Category: "
            + ReportService._filter_label(
                report_filter
            )
        )

        filter_cell.font = Font(
            size=10,
            bold=True,
            color=ReportService.GREEN,
        )

        filter_cell.alignment = Alignment(
            horizontal="center",
        )

        # ======================================================
        # TARGET INFORMATION
        # ======================================================

        worksheet.merge_cells(
            "A4:E4"
        )

        worksheet["A4"] = (
            "Vaccination Target: "
            f"{summary.get('vaccination_target', 83.0)}%"
        )

        worksheet["A4"].font = Font(
            bold=True,
            color=ReportService.DARK,
        )

        worksheet.merge_cells(
            "F4:K4"
        )

        worksheet["F4"] = (
            "Pashudhan Target: "
            f"{summary.get('pashudhan_target', 97.0)}%"
        )

        worksheet["F4"].font = Font(
            bold=True,
            color=ReportService.DARK,
        )

        # ======================================================
        # DISTRICT SUMMARY
        # ======================================================

        worksheet.merge_cells(
            "A5:K5"
        )

        worksheet["A5"] = (
            "District Summary | "
            f"Population: "
            f"{summary.get('target', 0):,} | "
            f"Vaccinations: "
            f"{summary.get('vaccinations', 0):,} | "
            f"Vaccination Achievement: "
            f"{summary.get('vaccination_percentage', 0)}% | "
            f"Pashudhan Entries: "
            f"{summary.get('entries', 0):,} | "
            f"Pashudhan Achievement: "
            f"{summary.get('pashudhan_percentage', 0)}%"
        )

        worksheet["A5"].font = Font(
            size=9,
            color=ReportService.GREY,
        )

        worksheet["A5"].alignment = Alignment(
            horizontal="center",
        )

        # ======================================================
        # GENERATED INFORMATION
        # ======================================================

        worksheet.merge_cells(
            "A6:K6"
        )

        generated_at = datetime.now().strftime(
            "%d-%m-%Y %I:%M %p"
        )

        worksheet["A6"] = (
            f"Generated: {generated_at} | "
            f"Records: {len(filtered_rows)}"
        )

        worksheet["A6"].font = Font(
            size=9,
            italic=True,
            color=ReportService.GREY,
        )

        worksheet["A6"].alignment = Alignment(
            horizontal="right",
        )

        # ======================================================
        # TABLE HEADER
        # ======================================================

        headers = [
            "Sl. No.",
            "Panchayath",
            "Total Squads",
            "Submitted",
            "Pending",
            "Population Target",
            "Vaccinations",
            "Vaccination %",
            "Pashudhan Entries",
            "Pashudhan %",
            "Status",
        ]

        header_row = 7

        for column_number, heading in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_number,
                value=heading,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.WHITE,
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.GREEN,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[
            header_row
        ].height = 30

        # ======================================================
        # TABLE DATA
        # ======================================================

        thin_border = Border(
            left=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            right=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            top=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            bottom=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
        )

        for serial_number, row in enumerate(
            filtered_rows,
            start=1,
        ):

            excel_row = (
                header_row
                + serial_number
            )

            values = [
                serial_number,
                row["panchayath"].name,
                row["total"],
                row["submitted"],
                row["pending"],
                row["target"],
                row["vaccinations"],
                row["vaccination_percentage"] / 100,
                row["entries"],
                row["pashudhan_percentage"] / 100,
                row["status"],
            ]

            for column_number, value in enumerate(
                values,
                start=1,
            ):

                cell = worksheet.cell(
                    row=excel_row,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center",
                    horizontal=(
                        "left"
                        if column_number == 2
                        else "center"
                    ),
                )

            # --------------------------------------------------
            # Percentage formatting
            # --------------------------------------------------

            worksheet.cell(
                row=excel_row,
                column=8,
            ).number_format = "0.00%"

            worksheet.cell(
                row=excel_row,
                column=10,
            ).number_format = "0.00%"

            # --------------------------------------------------
            # Panchayath Name
            # --------------------------------------------------

            worksheet.cell(
                row=excel_row,
                column=2,
            ).font = Font(
                bold=True,
                color=ReportService.DARK,
            )

            # --------------------------------------------------
            # Achievement Colours
            # --------------------------------------------------

            vaccination_cell = worksheet.cell(
                row=excel_row,
                column=8,
            )

            pashudhan_cell = worksheet.cell(
                row=excel_row,
                column=10,
            )

            if row["vaccination_target_met"]:

                vaccination_cell.font = Font(
                    bold=True,
                    color=ReportService.GREEN,
                )

            else:

                vaccination_cell.font = Font(
                    bold=True,
                    color=ReportService.RED,
                )

            if row["pashudhan_target_met"]:

                pashudhan_cell.font = Font(
                    bold=True,
                    color=ReportService.GREEN,
                )

            else:

                pashudhan_cell.font = Font(
                    bold=True,
                    color=ReportService.RED,
                )

            # --------------------------------------------------
            # Status Colour
            # --------------------------------------------------

            status_cell = worksheet.cell(
                row=excel_row,
                column=11,
            )

            ReportService._style_status_cell(
                status_cell,
                row["status"],
            )

        # ======================================================
        # COLUMN WIDTHS
        # ======================================================

        widths = {
            1: 10,
            2: 28,
            3: 13,
            4: 12,
            5: 10,
            6: 18,
            7: 16,
            8: 16,
            9: 18,
            10: 15,
            11: 20,
        }

        for column_number, width in widths.items():

            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = width

        # ======================================================
        # SAVE TO MEMORY
        # ======================================================

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output

    # ==========================================================
    # Panchayath Filter Matching
    # ==========================================================

    @staticmethod
    def _matches_filter(
        row,
        report_filter,
    ):

        if report_filter == "all":
            return True

        if report_filter == "vaccination-achieved":

            return bool(
                row["vaccination_target_met"]
            )

        if report_filter == "pashudhan-achieved":

            return bool(
                row["pashudhan_target_met"]
            )

        if report_filter == "both-achieved":

            return (
                bool(
                    row["vaccination_target_met"]
                )
                and bool(
                    row["pashudhan_target_met"]
                )
            )

        if report_filter == "low-vaccination":

            return not bool(
                row["vaccination_target_met"]
            )

        if report_filter == "low-pashudhan":

            return not bool(
                row["pashudhan_target_met"]
            )

        if report_filter == "low-both":

            return (
                not bool(
                    row["vaccination_target_met"]
                )
                and not bool(
                    row["pashudhan_target_met"]
                )
            )

        if report_filter == "pending":

            return row["status"] in (
                "Pending",
                "Not Started",
            )

        return True

    # ==========================================================
    # Panchayath Filter Display Label
    # ==========================================================

    @staticmethod
    def _filter_label(
        report_filter,
    ):

        labels = {
            "all":
                "All Panchayaths",

            "vaccination-achieved":
                "Vaccination Target Achieved",

            "pashudhan-achieved":
                "Pashudhan Target Achieved",

            "both-achieved":
                "Both Targets Achieved",

            "low-vaccination":
                "Low Vaccination",

            "low-pashudhan":
                "Low Pashudhan",

            "low-both":
                "Low Vaccination & Pashudhan",

            "pending":
                "Pending / Not Started",
        }

        return labels.get(
            report_filter,
            "All Panchayaths",
        )

    # ==========================================================
    # Panchayath Status Cell Styling
    # ==========================================================

    @staticmethod
    def _style_status_cell(
        cell,
        status,
    ):

        cell.font = Font(
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        if status == "Target Achieved":

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.LIGHT_GREEN,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.GREEN,
            )

            return

        if status in (
            "Pending",
            "Not Started",
        ):

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.LIGHT_AMBER,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.AMBER,
            )

            return

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=ReportService.LIGHT_RED,
        )

        cell.font = Font(
            bold=True,
            color=ReportService.RED,
        )

    # ==========================================================
    # Squad-wise Achievement Export
    # ==========================================================

    @staticmethod
    def create_squad_wise_workbook(
        campaign,
        monitoring,
        summary,
        report_filter="all",
    ):
        """
        Create the NCMS Squad-wise Excel report.

        Includes submitted and pending squads.

        Achievement filters apply only to submitted squads.
        Pending squads are never classified as low achievers.

        Supported filters:

        all
        submitted
        pending
        vaccination-achieved
        pashudhan-achieved
        both-achieved
        low-vaccination
        low-pashudhan
        low-both
        """

        # ------------------------------------------------------
        # Flatten Panchayath monitoring data into squads
        # ------------------------------------------------------

        squad_rows = []

        for panchayath_row in monitoring:

            panchayath = panchayath_row["panchayath"]

            # --------------------------------------------------
            # Submitted squads
            # --------------------------------------------------

            for squad in panchayath_row.get(
                "submitted_squads",
                [],
            ):

                squad_row = {
                    "panchayath": panchayath,
                    "submission_status": "submitted",
                    "squad": squad,
                }

                if ReportService._matches_squad_filter(
                    squad_row,
                    report_filter,
                ):
                    squad_rows.append(
                        squad_row
                    )

            # --------------------------------------------------
            # Pending squads
            # --------------------------------------------------

            for squad in panchayath_row.get(
                "pending_squads",
                [],
            ):

                squad_row = {
                    "panchayath": panchayath,
                    "submission_status": "pending",
                    "squad": squad,
                }

                if ReportService._matches_squad_filter(
                    squad_row,
                    report_filter,
                ):
                    squad_rows.append(
                        squad_row
                    )

        # ------------------------------------------------------
        # Workbook
        # ------------------------------------------------------

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Squad-wise Report"

        worksheet.sheet_view.showGridLines = False

        worksheet.freeze_panes = "A8"

        worksheet.page_setup.orientation = "landscape"

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        worksheet.print_title_rows = "1:7"

        # ======================================================
        # TITLE
        # ======================================================

        worksheet.merge_cells(
            "A1:P1"
        )

        title_cell = worksheet["A1"]

        title_cell.value = (
            "NCMS - Squad-wise Campaign Report"
        )

        title_cell.font = Font(
            size=16,
            bold=True,
            color=ReportService.WHITE,
        )

        title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=ReportService.GREEN,
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 28

        # ======================================================
        # CAMPAIGN
        # ======================================================

        worksheet.merge_cells(
            "A2:P2"
        )

        campaign_name = (
            campaign.name
            if campaign
            else ""
        )

        worksheet["A2"] = (
            f"Campaign: {campaign_name}"
        )

        worksheet["A2"].font = Font(
            size=11,
            bold=True,
            color=ReportService.DARK,
        )

        worksheet["A2"].alignment = Alignment(
            horizontal="center",
        )

        # ======================================================
        # REPORT CATEGORY
        # ======================================================

        worksheet.merge_cells(
            "A3:P3"
        )

        worksheet["A3"] = (
            "Report Category: "
            + ReportService._squad_filter_label(
                report_filter
            )
        )

        worksheet["A3"].font = Font(
            size=10,
            bold=True,
            color=ReportService.GREEN,
        )

        worksheet["A3"].alignment = Alignment(
            horizontal="center",
        )

        # ======================================================
        # TARGETS
        # ======================================================

        worksheet.merge_cells(
            "A4:H4"
        )

        worksheet["A4"] = (
            "Vaccination Target: "
            f"{summary.get('vaccination_target', 83.0)}%"
        )

        worksheet["A4"].font = Font(
            bold=True,
            color=ReportService.DARK,
        )

        worksheet.merge_cells(
            "I4:P4"
        )

        worksheet["I4"] = (
            "Pashudhan Target: "
            f"{summary.get('pashudhan_target', 97.0)}%"
        )

        worksheet["I4"].font = Font(
            bold=True,
            color=ReportService.DARK,
        )

        # ======================================================
        # SUBMISSION SUMMARY
        # ======================================================

        worksheet.merge_cells(
            "A5:P5"
        )

        worksheet["A5"] = (
            "Campaign Summary | "
            f"Total Squads: "
            f"{summary.get('total_squads', 0)} | "
            f"Submitted: "
            f"{summary.get('submitted_squads', 0)} | "
            f"Pending: "
            f"{summary.get('pending_squads', 0)} | "
            f"Submission Completion: "
            f"{summary.get('submission_percentage', 0)}%"
        )

        worksheet["A5"].font = Font(
            size=9,
            color=ReportService.GREY,
        )

        worksheet["A5"].alignment = Alignment(
            horizontal="center",
        )

        # ======================================================
        # GENERATED INFORMATION
        # ======================================================

        worksheet.merge_cells(
            "A6:P6"
        )

        generated_at = datetime.now().strftime(
            "%d-%m-%Y %I:%M %p"
        )

        worksheet["A6"] = (
            f"Generated: {generated_at} | "
            f"Squads: {len(squad_rows)}"
        )

        worksheet["A6"].font = Font(
            size=9,
            italic=True,
            color=ReportService.GREY,
        )

        worksheet["A6"].alignment = Alignment(
            horizontal="right",
        )

        # ======================================================
        # TABLE HEADER
        # ======================================================

        headers = [
            "Sl. No.",
            "Panchayath",
            "Squad No.",
            "Member Name",
            "Office",
            "Pashudhan ID",
            "Target",
            "Vaccinated",
            "Vaccination %",
            "Pashudhan Entries",
            "Pashudhan %",
            "Status",
            "Vaccination Reason",
            "Pashudhan Reason",
            "Submission Ref.",
            "Submitted Date",
        ]

        header_row = 7

        for column_number, heading in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_number,
                value=heading,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.WHITE,
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.GREEN,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[
            header_row
        ].height = 32

        # ======================================================
        # BORDER
        # ======================================================

        thin_border = Border(
            left=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            right=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            top=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            bottom=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
        )

        # ======================================================
        # DATA
        # ======================================================

        for serial_number, item in enumerate(
            squad_rows,
            start=1,
        ):

            excel_row = (
                header_row
                + serial_number
            )

            panchayath = item[
                "panchayath"
            ]

            squad = item[
                "squad"
            ]

            is_submitted = (
                item["submission_status"]
                == "submitted"
            )

            # --------------------------------------------------
            # Members
            # --------------------------------------------------

            members = squad.get(
                "members",
                [],
            )

            member_names = []
            offices = []
            pashudhan_ids = []

            for member in members:

                member_name = (
                    member.get(
                        "member_name"
                    )
                    or ""
                )

                office = (
                    member.get(
                        "office"
                    )
                    or ""
                )

                pashudhan_id = (
                    member.get(
                        "pashudhan_id"
                    )
                    or ""
                )

                if member_name:

                    member_names.append(
                        str(member_name)
                    )

                if office:

                    offices.append(
                        str(office)
                    )

                if pashudhan_id:

                    pashudhan_ids.append(
                        str(pashudhan_id)
                    )

            member_text = "\n".join(
                member_names
            )

            office_text = "\n".join(
                offices
            )

            pashudhan_text = "\n".join(
                pashudhan_ids
            )

            # --------------------------------------------------
            # Submitted squad
            # --------------------------------------------------

            if is_submitted:

                vaccination_met = bool(
                    squad.get(
                        "vaccination_target_met",
                        False,
                    )
                )

                pashudhan_met = bool(
                    squad.get(
                        "pashudhan_target_met",
                        False,
                    )
                )

                if (
                    vaccination_met
                    and pashudhan_met
                ):

                    status = (
                        "Both Achieved"
                    )

                elif (
                    not vaccination_met
                    and not pashudhan_met
                ):

                    status = (
                        "Low Both"
                    )

                elif not vaccination_met:

                    status = (
                        "Low Vaccination"
                    )

                else:

                    status = (
                        "Low Pashudhan"
                    )

                vaccination_percentage = (
                    squad.get(
                        "vaccination_percentage",
                        0,
                    )
                    / 100
                )

                pashudhan_percentage = (
                    squad.get(
                        "pashudhan_percentage",
                        0,
                    )
                    / 100
                )

                submitted_at = squad.get(
                    "submitted_at"
                )

                if submitted_at:

                    if hasattr(
                        submitted_at,
                        "strftime",
                    ):

                        submitted_at_value = (
                            submitted_at.strftime(
                                "%d-%m-%Y %I:%M %p"
                            )
                        )

                    else:

                        submitted_at_value = str(
                            submitted_at
                        )

                else:

                    submitted_at_value = ""

                values = [
                    serial_number,
                    panchayath.name,
                    squad.get(
                        "squad_no",
                        "",
                    ),
                    member_text,
                    office_text,
                    pashudhan_text,
                    squad.get(
                        "target",
                        0,
                    ),
                    squad.get(
                        "vaccinations",
                        0,
                    ),
                    vaccination_percentage,
                    squad.get(
                        "entries",
                        0,
                    ),
                    pashudhan_percentage,
                    status,
                    squad.get(
                        "vaccination_reason",
                        "",
                    ) or "",
                    squad.get(
                        "pashudhan_reason",
                        "",
                    ) or "",
                    squad.get(
                        "submission_token",
                        "",
                    ) or "",
                    submitted_at_value,
                ]

            # --------------------------------------------------
            # Pending squad
            # --------------------------------------------------

            else:

                vaccination_met = False
                pashudhan_met = False

                status = "Pending"

                values = [
                    serial_number,
                    panchayath.name,
                    squad.get(
                        "squad_no",
                        "",
                    ),
                    member_text,
                    office_text,
                    pashudhan_text,
                    "",
                    "",
                    "",
                    "",
                    "",
                    status,
                    "",
                    "",
                    "",
                    "",
                ]

            # --------------------------------------------------
            # Write row
            # --------------------------------------------------

            for column_number, value in enumerate(
                values,
                start=1,
            ):

                cell = worksheet.cell(
                    row=excel_row,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border

                cell.alignment = Alignment(
                    horizontal=(
                        "left"
                        if column_number in (
                            2,
                            4,
                            5,
                            6,
                            13,
                            14,
                            15,
                        )
                        else "center"
                    ),
                    vertical="top",
                    wrap_text=True,
                )

            # --------------------------------------------------
            # Panchayath / Squad styling
            # --------------------------------------------------

            worksheet.cell(
                row=excel_row,
                column=2,
            ).font = Font(
                bold=True,
                color=ReportService.DARK,
            )

            worksheet.cell(
                row=excel_row,
                column=3,
            ).font = Font(
                bold=True,
                color=ReportService.DARK,
            )

            # --------------------------------------------------
            # Percentage styling
            # --------------------------------------------------

            if is_submitted:

                vaccination_cell = (
                    worksheet.cell(
                        row=excel_row,
                        column=9,
                    )
                )

                pashudhan_cell = (
                    worksheet.cell(
                        row=excel_row,
                        column=11,
                    )
                )

                vaccination_cell.number_format = (
                    "0.00%"
                )

                pashudhan_cell.number_format = (
                    "0.00%"
                )

                vaccination_cell.font = Font(
                    bold=True,
                    color=(
                        ReportService.GREEN
                        if vaccination_met
                        else ReportService.RED
                    ),
                )

                pashudhan_cell.font = Font(
                    bold=True,
                    color=(
                        ReportService.GREEN
                        if pashudhan_met
                        else ReportService.RED
                    ),
                )

            # --------------------------------------------------
            # Status styling
            # --------------------------------------------------

            status_cell = worksheet.cell(
                row=excel_row,
                column=12,
            )

            ReportService._style_squad_status_cell(
                status_cell,
                status,
            )

        # ======================================================
        # EXCEL AUTO FILTER
        # ======================================================

        last_row = max(
            header_row,
            header_row + len(
                squad_rows
            ),
        )

        worksheet.auto_filter.ref = (
            f"A{header_row}:P{last_row}"
        )

        # ======================================================
        # COLUMN WIDTHS
        # ======================================================

        widths = {
            1: 9,
            2: 25,
            3: 11,
            4: 25,
            5: 28,
            6: 18,
            7: 12,
            8: 13,
            9: 15,
            10: 18,
            11: 14,
            12: 18,
            13: 32,
            14: 32,
            15: 22,
            16: 21,
        }

        for column_number, width in widths.items():

            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = width

        # ======================================================
        # SAVE TO MEMORY
        # ======================================================

        output = BytesIO()

        workbook.save(
            output
        )

        output.seek(0)

        return output

    # ==========================================================
    # Squad Filter Matching
    # ==========================================================

    @staticmethod
    def _matches_squad_filter(
        item,
        report_filter,
    ):

        submission_status = item.get(
            "submission_status",
            ""
        )

        squad = item.get(
            "squad",
            {}
        )

        if report_filter == "all":
            return True

        if report_filter == "submitted":

            return (
                submission_status
                == "submitted"
            )

        if report_filter == "pending":

            return (
                submission_status
                == "pending"
            )

        # ------------------------------------------------------
        # Achievement filters only apply to submitted squads.
        # Pending squads are not low achievers.
        # ------------------------------------------------------

        if submission_status != "submitted":
            return False

        vaccination_met = bool(
            squad.get(
                "vaccination_target_met",
                False,
            )
        )

        pashudhan_met = bool(
            squad.get(
                "pashudhan_target_met",
                False,
            )
        )

        if report_filter == "vaccination-achieved":

            return vaccination_met

        if report_filter == "pashudhan-achieved":

            return pashudhan_met

        if report_filter == "both-achieved":

            return (
                vaccination_met
                and pashudhan_met
            )

        if report_filter == "low-vaccination":

            return not vaccination_met

        if report_filter == "low-pashudhan":

            return not pashudhan_met

        if report_filter == "low-both":

            return (
                not vaccination_met
                and not pashudhan_met
            )

        return True

    # ==========================================================
    # Squad Filter Label
    # ==========================================================

    @staticmethod
    def _squad_filter_label(
        report_filter,
    ):

        labels = {
            "all":
                "All Squads",

            "submitted":
                "Submitted Squads",

            "pending":
                "Pending Squads",

            "vaccination-achieved":
                "Vaccination Target Achieved",

            "pashudhan-achieved":
                "Pashudhan Target Achieved",

            "both-achieved":
                "Both Targets Achieved",

            "low-vaccination":
                "Low Vaccination",

            "low-pashudhan":
                "Low Pashudhan",

            "low-both":
                "Low Vaccination & Pashudhan",
        }

        return labels.get(
            report_filter,
            "All Squads",
        )

    # ==========================================================
    # Squad Status Cell Styling
    # ==========================================================

    @staticmethod
    def _style_squad_status_cell(
        cell,
        status,
    ):

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        if status == "Both Achieved":

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.LIGHT_GREEN,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.GREEN,
            )

            return

        if status == "Pending":

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.LIGHT_AMBER,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.AMBER,
            )

            return

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=ReportService.LIGHT_RED,
        )

        cell.font = Font(
            bold=True,
            color=ReportService.RED,
        )

    # ==========================================================
    # Pending Submission Export
    # ==========================================================

    @staticmethod
    def create_pending_submissions_workbook(
        campaign,
        monitoring,
        summary,
    ):
        """
        Create the NCMS Pending Submission Excel report.

        One row represents one pending squad.
        Multiple squad members are retained within the same row.
        """

        # ------------------------------------------------------
        # Collect pending squads
        # ------------------------------------------------------

        pending_rows = []

        for panchayath_row in monitoring:

            panchayath = panchayath_row["panchayath"]

            for squad in panchayath_row.get(
                "pending_squads",
                [],
            ):

                pending_rows.append(
                    {
                        "panchayath": panchayath,
                        "squad": squad,
                    }
                )

        # ------------------------------------------------------
        # Workbook
        # ------------------------------------------------------

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Pending Submissions"

        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A7"

        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        worksheet.print_title_rows = "1:6"


        # ======================================================
        # TITLE
        # ======================================================

        worksheet.merge_cells("A1:G1")

        title_cell = worksheet["A1"]

        title_cell.value = (
            "NCMS - Pending Submission Report"
        )

        title_cell.font = Font(
            size=16,
            bold=True,
            color=ReportService.WHITE,
        )

        title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=ReportService.GREEN,
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 28


        # ======================================================
        # CAMPAIGN
        # ======================================================

        worksheet.merge_cells("A2:G2")

        campaign_name = (
            campaign.name
            if campaign
            else ""
        )

        worksheet["A2"] = (
            f"Campaign: {campaign_name}"
        )

        worksheet["A2"].font = Font(
            size=11,
            bold=True,
            color=ReportService.DARK,
        )

        worksheet["A2"].alignment = Alignment(
            horizontal="center",
        )


        # ======================================================
        # REPORT CATEGORY
        # ======================================================

        worksheet.merge_cells("A3:G3")

        worksheet["A3"] = (
            "Report Category: Pending Submissions"
        )

        worksheet["A3"].font = Font(
            size=10,
            bold=True,
            color=ReportService.AMBER,
        )

        worksheet["A3"].alignment = Alignment(
            horizontal="center",
        )


        # ======================================================
        # CAMPAIGN SUMMARY
        # ======================================================

        worksheet.merge_cells("A4:G4")

        worksheet["A4"] = (
            "Campaign Summary | "
            f"Total Squads: "
            f"{summary.get('total_squads', 0)} | "
            f"Submitted: "
            f"{summary.get('submitted_squads', 0)} | "
            f"Pending: "
            f"{summary.get('pending_squads', 0)} | "
            f"Submission Completion: "
            f"{summary.get('submission_percentage', 0)}%"
        )

        worksheet["A4"].font = Font(
            size=9,
            color=ReportService.GREY,
        )

        worksheet["A4"].alignment = Alignment(
            horizontal="center",
        )


        # ======================================================
        # GENERATED INFORMATION
        # ======================================================

        worksheet.merge_cells("A5:G5")

        generated_at = datetime.now().strftime(
            "%d-%m-%Y %I:%M %p"
        )

        worksheet["A5"] = (
            f"Generated: {generated_at} | "
            f"Pending Squads: {len(pending_rows)}"
        )

        worksheet["A5"].font = Font(
            size=9,
            italic=True,
            color=ReportService.GREY,
        )

        worksheet["A5"].alignment = Alignment(
            horizontal="right",
        )


        # ======================================================
        # TABLE HEADER
        # ======================================================

        headers = [
            "Sl. No.",
            "Panchayath",
            "Squad No.",
            "Member Name",
            "Office",
            "Pashudhan ID",
            "Status",
        ]

        header_row = 6

        for column_number, heading in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_number,
                value=heading,
            )

            cell.font = Font(
                bold=True,
                color=ReportService.WHITE,
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.GREEN,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[
            header_row
        ].height = 30


        # ======================================================
        # BORDER
        # ======================================================

        thin_border = Border(
            left=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            right=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            top=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
            bottom=Side(
                style="thin",
                color=ReportService.BORDER_GREY,
            ),
        )


        # ======================================================
        # DATA
        # ======================================================

        for serial_number, item in enumerate(
            pending_rows,
            start=1,
        ):

            excel_row = (
                header_row
                + serial_number
            )

            panchayath = item["panchayath"]
            squad = item["squad"]

            members = squad.get(
                "members",
                [],
            )

            member_names = []
            offices = []
            pashudhan_ids = []

            for member in members:

                member_name = (
                    member.get("member_name")
                    or ""
                )

                office = (
                    member.get("office")
                    or ""
                )

                pashudhan_id = (
                    member.get("pashudhan_id")
                    or ""
                )

                if member_name:
                    member_names.append(
                        str(member_name)
                    )

                if office:
                    offices.append(
                        str(office)
                    )

                if pashudhan_id:
                    pashudhan_ids.append(
                        str(pashudhan_id)
                    )

            member_text = "\n".join(
                member_names
            )

            office_text = "\n".join(
                offices
            )

            pashudhan_text = "\n".join(
                pashudhan_ids
            )

            values = [
                serial_number,
                panchayath.name,
                squad.get(
                    "squad_no",
                    "",
                ),
                member_text,
                office_text,
                pashudhan_text,
                "Pending",
            ]

            for column_number, value in enumerate(
                values,
                start=1,
            ):

                cell = worksheet.cell(
                    row=excel_row,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border

                cell.alignment = Alignment(
                    horizontal=(
                        "left"
                        if column_number in (
                            2,
                            4,
                            5,
                            6,
                        )
                        else "center"
                    ),
                    vertical="top",
                    wrap_text=True,
                )


            # --------------------------------------------------
            # Panchayath / Squad styling
            # --------------------------------------------------

            worksheet.cell(
                row=excel_row,
                column=2,
            ).font = Font(
                bold=True,
                color=ReportService.DARK,
            )

            worksheet.cell(
                row=excel_row,
                column=3,
            ).font = Font(
                bold=True,
                color=ReportService.DARK,
            )


            # --------------------------------------------------
            # Pending Status
            # --------------------------------------------------

            status_cell = worksheet.cell(
                row=excel_row,
                column=7,
            )

            status_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=ReportService.LIGHT_AMBER,
            )

            status_cell.font = Font(
                bold=True,
                color=ReportService.AMBER,
            )

            status_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


        # ======================================================
        # AUTO FILTER
        # ======================================================

        last_row = max(
            header_row,
            header_row + len(pending_rows),
        )

        worksheet.auto_filter.ref = (
            f"A{header_row}:G{last_row}"
        )


        # ======================================================
        # COLUMN WIDTHS
        # ======================================================

        widths = {
            1: 9,
            2: 28,
            3: 12,
            4: 28,
            5: 32,
            6: 20,
            7: 14,
        }

        for column_number, width in widths.items():

            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = width


        # ======================================================
        # SAVE TO MEMORY
        # ======================================================

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        return output
